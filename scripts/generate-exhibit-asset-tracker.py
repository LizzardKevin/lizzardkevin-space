#!/usr/bin/env python3
"""Generate exhibit-asset-tracker.xlsx from embedded sheet data."""

from __future__ import annotations

import csv
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
ASSETS = ROOT / "docs" / "assets"

EXHIBITS_HEADERS = [
    "exhibitId",
    "sceneObjectName",
    "type",
    "zone",
    "status",
    "in_space_main",
    "folder_path",
    "focus_glb_file",
    "focus_glb_exists",
    "audio_file",
    "audio_exists",
    "video_file",
    "video_exists",
    "image_file",
    "image_exists",
    "in_manifest",
    "manifest_buttons",
    "focusGlbUrl",
    "notes",
    "last_updated",
    "owner",
]

EXHIBITS_ROWS = []

SCENE_HEADERS = ["asset_key", "path", "exists", "status", "notes"]

SCENE_ROWS = [
    [
        "space_main",
        "models/space_main.glb",
        "Y",
        "in_library",
        "Production gallery space with COL_floor_* helpers, bulb_* lights, and spawn_player_main.",
    ],
    [
        "gallery_main_demo",
        "models/gallery_main.glb",
        "Y",
        "historical_demo",
        "Historical comparison asset; not used as the production SPACE scene.",
    ],
    [
        "bgm_architecture",
        "audio/bgm_architecture.mp3",
        "N",
        "todo",
        "Optional architecture-zone BGM after entering SPACE.",
    ],
    ["wall_art", "media/art_01.jpg", "N", "optional", "Optional north-wall code-art texture."],
]

LEGEND_ROWS = [
    ["field/value", "description"],
    ["status: todo", "Not started."],
    ["status: in_progress", "DCC work or export is in progress."],
    ["status: in_library", "File is present under the public runtime path."],
    ["status: needs_adjustment", "File exists but naming, scale, or manifest data needs work."],
    ["status: accepted", "Runtime integration has been checked."],
    ["Y / N", "Whether the file or binding is present."],
    ["exhibitId", "Must match manifest.json exhibitId."],
    ["sceneObjectName", "space_main hit mesh name; convention is exhibit_<exhibitId>."],
    ["focus_glb_file", "Stored under folder_path; convention is focus_<exhibitId>.glb."],
    ["in_manifest", "Whether manifest.json has this item."],
    ["maintenance order", "Add row, create folder, add media/model, update manifest, then mark ready."],
    ["reference docs", "docs/gallery-mesh-naming.md / docs/asset-manifest.md"],
]


def write_csv(path: Path, headers: list[str], rows: list[list[str]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)


def write_xlsx(path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        raise SystemExit(
            "openpyxl is required to build .xlsx. Install with: pip install openpyxl"
        ) from exc

    wb = Workbook()
    header_font = Font(bold=True)

    ws1 = wb.active
    ws1.title = "exhibits"
    ws1.append(EXHIBITS_HEADERS)
    for cell in ws1[1]:
        cell.font = header_font
    for row in EXHIBITS_ROWS:
        ws1.append(row)

    ws2 = wb.create_sheet("scene_assets")
    ws2.append(SCENE_HEADERS)
    for cell in ws2[1]:
        cell.font = header_font
    for row in SCENE_ROWS:
        ws2.append(row)

    ws3 = wb.create_sheet("legend")
    for row in LEGEND_ROWS:
        ws3.append(row)
    for cell in ws3[1]:
        cell.font = header_font

    wb.save(path)


def main() -> None:
    ASSETS.mkdir(parents=True, exist_ok=True)

    write_csv(ASSETS / "exhibit-asset-tracker-exhibits.csv", EXHIBITS_HEADERS, EXHIBITS_ROWS)
    write_csv(ASSETS / "exhibit-asset-tracker-scene_assets.csv", SCENE_HEADERS, SCENE_ROWS)
    write_csv(ASSETS / "exhibit-asset-tracker-legend.csv", LEGEND_ROWS[0], LEGEND_ROWS[1:])

    write_xlsx(ASSETS / "exhibit-asset-tracker.xlsx")
    print(f"Wrote tracker files under {ASSETS}")


if __name__ == "__main__":
    main()
